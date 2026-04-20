"""
Excel export voor het Werkzaamheden Weekrapport.
Gebruikt openpyxl om een opgemaakt .xlsx bestand te genereren.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

DAYS = ["ma", "di", "wo", "do", "vr", "za", "zo"]
DAYS_NL = {
    "ma": "Maandag", "di": "Dinsdag", "wo": "Woensdag",
    "do": "Donderdag", "vr": "Vrijdag", "za": "Zaterdag", "zo": "Zondag"
}
DAYS_SHORT = {
    "ma": "Ma", "di": "Di", "wo": "Wo",
    "do": "Do", "vr": "Vr", "za": "Za", "zo": "Zo"
}

HEADER_COLOR = "343A40"
HEADER_ALT_COLOR = "495057"
SUBHEADER_COLOR = "6C757D"
GREEN_COLOR = "D4EDDA"
LIGHT_COLOR = "F8F9FA"
WHITE_COLOR = "FFFFFF"
ACCENT_COLOR = "0D6EFD"


def _header_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _header_font(color: str = "FFFFFF", bold: bool = True, size: int = 10) -> Font:
    return Font(name="Segoe UI", bold=bold, color=color, size=size)


def _normal_font(bold: bool = False, size: int = 10) -> Font:
    return Font(name="Segoe UI", bold=bold, size=size)


def _border_thin() -> Border:
    thin = Side(style="thin", color="DEE2E6")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _apply_row_style(ws, row: int, cols: int, fill_color: str,
                     font_color: str = "212529", bold: bool = False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _header_fill(fill_color)
        cell.font = Font(name="Segoe UI", bold=bold, color=font_color, size=10)
        cell.border = _border_thin()
        cell.alignment = _center()


def _write_week_to_sheet(ws, data: dict) -> None:
    """Schrijft één weekrapport naar een bestaand werkblad."""
    project_nr = data.get("project_nr", "")
    project_start = data.get("project_start", "")
    project_end = data.get("project_end", "")
    project_week_nr = data.get("project_week_nr", "")
    kalender_week = data.get("kalender_week_nr", "")
    week_start = data.get("week_start", "")
    week_end = data.get("week_end", "")
    locatie = data.get("locatie", "")

    # Datumopmaak: 2025-04-07 → 07/04/2025
    def fmt_date(d: str) -> str:
        try:
            from datetime import date
            dt = date.fromisoformat(d)
            return dt.strftime("%d/%m/%Y")
        except Exception:
            return d

    # Kolombreedte instellen
    col_widths = [18, 20, 16, 12, 10, 10, 10, 10, 10, 10, 10, 10, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    current_row = 1

    # ─── TITEL ───────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{current_row}:M{current_row}")
    title_cell = ws.cell(row=current_row, column=1, value="Werkzaamheden Rapport")
    title_cell.fill = _header_fill(HEADER_COLOR)
    title_cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=16)
    title_cell.alignment = _center()
    ws.row_dimensions[current_row].height = 32
    current_row += 1

    # ─── PROJECTINFO ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A{current_row}:C{current_row}")
    ws.cell(row=current_row, column=1, value=f"Project: {project_nr}").font = _normal_font(bold=True)
    ws.cell(row=current_row, column=1).alignment = _left()

    ws.cell(row=current_row, column=4, value=f"Weeknr: {project_week_nr}").font = _normal_font(bold=True)
    ws.cell(row=current_row, column=4).alignment = _center()

    ws.merge_cells(f"E{current_row}:G{current_row}")
    ws.cell(row=current_row, column=5, value=f"Kalender week: {kalender_week}").font = _normal_font(bold=True)
    ws.cell(row=current_row, column=5).alignment = _center()

    ws.merge_cells(f"H{current_row}:J{current_row}")
    week_range = f"Week: {fmt_date(week_start)} - {fmt_date(week_end)}"
    ws.cell(row=current_row, column=8, value=week_range).font = _normal_font()
    ws.cell(row=current_row, column=8).alignment = _center()

    if project_start or project_end:
        ws.merge_cells(f"K{current_row}:M{current_row}")
        info = f"Project: {project_start or '?'} – {project_end or '?'}"
        ws.cell(row=current_row, column=11, value=info).font = _normal_font(bold=True)
        ws.cell(row=current_row, column=11).alignment = _center()

    for c in range(1, 14):
        ws.cell(row=current_row, column=c).fill = _header_fill(LIGHT_COLOR)
        ws.cell(row=current_row, column=c).border = _border_thin()
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    if locatie:
        ws.merge_cells(f"A{current_row}:M{current_row}")
        ws.cell(row=current_row, column=1, value=f"Locatie: {locatie}").font = _normal_font()
        ws.cell(row=current_row, column=1).alignment = _left()
        ws.cell(row=current_row, column=1).fill = _header_fill(LIGHT_COLOR)
        ws.cell(row=current_row, column=1).border = _border_thin()
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    current_row += 1  # lege rij

    # ─── WEER RAPPORT ─────────────────────────────────────────────────────────
    ws.merge_cells(f"A{current_row}:E{current_row}")
    ws.cell(row=current_row, column=1, value="Weer Rapport").fill = _header_fill(HEADER_COLOR)
    ws.cell(row=current_row, column=1).font = _header_font(size=12)
    ws.cell(row=current_row, column=1).alignment = _center()
    for c in range(2, 6):
        ws.cell(row=current_row, column=c).fill = _header_fill(HEADER_COLOR)
        ws.cell(row=current_row, column=c).border = _border_thin()
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    # Kolomkoppen weer
    weer_headers = ["Dag", "Weer", "Temperatuur (°C)", "Regen (mm)", "Windkracht (Bft)"]
    for c, h in enumerate(weer_headers, 1):
        cell = ws.cell(row=current_row, column=c, value=h)
        cell.fill = _header_fill(HEADER_ALT_COLOR)
        cell.font = _header_font()
        cell.alignment = _center()
        cell.border = _border_thin()
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    # Weersdata
    weer = data.get("weer", {})
    for i, dag in enumerate(DAYS):
        w = weer.get(dag, {})
        row_color = WHITE_COLOR if i % 2 == 0 else LIGHT_COLOR
        vals = [
            DAYS_NL.get(dag, dag),
            w.get("beschrijving", ""),
            w.get("temp_c", ""),
            w.get("regen_mm", ""),
            w.get("wind_bft", "")
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=current_row, column=c, value=val if val is not None else "")
            cell.fill = _header_fill(row_color)
            cell.font = _normal_font()
            cell.alignment = _center()
            cell.border = _border_thin()
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    current_row += 1  # lege rij

    # ─── WERKBAARHEID ─────────────────────────────────────────────────────────
    ws.merge_cells(f"A{current_row}:E{current_row}")
    ws.cell(row=current_row, column=1, value="Werkbaarheid").fill = _header_fill(HEADER_COLOR)
    ws.cell(row=current_row, column=1).font = _header_font(size=12)
    ws.cell(row=current_row, column=1).alignment = _center()
    for c in range(2, 6):
        ws.cell(row=current_row, column=c).fill = _header_fill(HEADER_COLOR)
        ws.cell(row=current_row, column=c).border = _border_thin()
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    wb_data = data.get("werkbaarheid", {})
    feestdagen = wb_data.get("onwerkbaar_feestdagen", 0) or 0
    onwerkbaar_weer = wb_data.get("onwerkbaar_weer", 0) or 0
    werkbaar_vals = [
        ("Werkbare dagen (Ma-Vr)",           wb_data.get("werkbare_dagen", 0), GREEN_COLOR),
        ("Totaal onwerkbare dagen (Ma-Vr)",  feestdagen + onwerkbaar_weer,     "F8D7DA"),
        ("  w.v. Feestdagen",                feestdagen,                       "FFF3CD"),
        ("  w.v. Weersomstandigheden",       onwerkbaar_weer,                  "FFE8B0"),
    ]
    for label, val, kleur in werkbaar_vals:
        ws.merge_cells(f"A{current_row}:D{current_row}")
        cell_l = ws.cell(row=current_row, column=1, value=label)
        cell_l.font = _normal_font(bold=True)
        cell_l.alignment = _left()
        cell_l.fill = _header_fill(LIGHT_COLOR)
        cell_l.border = _border_thin()
        cell_v = ws.cell(row=current_row, column=5, value=val)
        cell_v.font = _normal_font(bold=True)
        cell_v.alignment = _center()
        cell_v.fill = _header_fill(kleur)
        cell_v.border = _border_thin()
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    current_row += 1  # lege rij

    # ─── UREN WERKNEMERS ──────────────────────────────────────────────────────
    num_cols = 13  # Naam | Bedrijf | Functie | Ma..Zo | Totaal
    ws.merge_cells(f"A{current_row}:{get_column_letter(num_cols)}{current_row}")
    ws.cell(row=current_row, column=1, value="Uren Werknemers").fill = _header_fill(HEADER_COLOR)
    ws.cell(row=current_row, column=1).font = _header_font(size=12)
    ws.cell(row=current_row, column=1).alignment = _center()
    for c in range(2, num_cols + 1):
        ws.cell(row=current_row, column=c).fill = _header_fill(HEADER_COLOR)
        ws.cell(row=current_row, column=c).border = _border_thin()
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    wn_headers = ["Naam Werknemer", "Naam Bedrijf", "Functie"] + \
                 [f"Uren {DAYS_SHORT[d]}" for d in DAYS] + ["Totaal"]
    for c, h in enumerate(wn_headers, 1):
        cell = ws.cell(row=current_row, column=c, value=h)
        cell.fill = _header_fill(HEADER_ALT_COLOR)
        cell.font = _header_font()
        cell.alignment = _center()
        cell.border = _border_thin()
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    werknemers = data.get("werknemers", [])
    dag_totalen = {d: 0.0 for d in DAYS}
    for i, wn in enumerate(werknemers):
        row_color = WHITE_COLOR if i % 2 == 0 else LIGHT_COLOR
        uren = wn.get("uren", {})
        totaal = sum(uren.get(d, 0) or 0 for d in DAYS)
        vals = (
            [wn.get("naam", ""), wn.get("bedrijf", ""), wn.get("functie", "")] +
            [uren.get(d, 0) or 0 for d in DAYS] +
            [totaal]
        )
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=current_row, column=c, value=val)
            cell.fill = _header_fill(row_color)
            cell.font = _normal_font()
            cell.alignment = _center() if c > 3 else _left()
            cell.border = _border_thin()
        ws.row_dimensions[current_row].height = 18
        current_row += 1
        for d in DAYS:
            dag_totalen[d] += uren.get(d, 0) or 0

    # Totaalrij
    totaal_alle = sum(dag_totalen[d] for d in DAYS)
    totaal_vals = ["Totaal", "", ""] + [dag_totalen[d] for d in DAYS] + [totaal_alle]
    for c, val in enumerate(totaal_vals, 1):
        cell = ws.cell(row=current_row, column=c, value=val if val != 0.0 or c > 3 else "")
        cell.fill = _header_fill(HEADER_ALT_COLOR)
        cell.font = _header_font()
        cell.alignment = _center() if c > 3 else _left()
        cell.border = _border_thin()
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    current_row += 1  # lege rij

    # ─── WERKZAAMHEDEN ────────────────────────────────────────────────────────
    ws.merge_cells(f"A{current_row}:M{current_row}")
    ws.cell(row=current_row, column=1, value="Werkzaamheden").fill = _header_fill(HEADER_COLOR)
    ws.cell(row=current_row, column=1).font = _header_font(size=12)
    ws.cell(row=current_row, column=1).alignment = _center()
    for c in range(2, 14):
        ws.cell(row=current_row, column=c).fill = _header_fill(HEADER_COLOR)
        ws.cell(row=current_row, column=c).border = _border_thin()
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    werkzaamheden = data.get("werkzaamheden", {})
    for dag in DAYS:
        activiteiten = [a for a in werkzaamheden.get(dag, []) if a.strip()]
        if not activiteiten and dag in ("za", "zo"):
            continue  # Sla lege Za/Zo over in export

        # Dag-header
        ws.merge_cells(f"A{current_row}:M{current_row}")
        ws.cell(row=current_row, column=1, value=DAYS_NL.get(dag, dag)).fill = _header_fill(HEADER_ALT_COLOR)
        ws.cell(row=current_row, column=1).font = _header_font()
        ws.cell(row=current_row, column=1).alignment = _left()
        for c in range(2, 14):
            ws.cell(row=current_row, column=c).fill = _header_fill(HEADER_ALT_COLOR)
            ws.cell(row=current_row, column=c).border = _border_thin()
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        if activiteiten:
            for j, activiteit in enumerate(activiteiten):
                row_color = WHITE_COLOR if j % 2 == 0 else LIGHT_COLOR
                ws.merge_cells(f"A{current_row}:M{current_row}")
                cell = ws.cell(row=current_row, column=1, value=activiteit)
                cell.fill = _header_fill(row_color)
                cell.font = _normal_font()
                cell.alignment = _left()
                cell.border = _border_thin()
                for c in range(2, 14):
                    ws.cell(row=current_row, column=c).fill = _header_fill(row_color)
                    ws.cell(row=current_row, column=c).border = _border_thin()
                ws.row_dimensions[current_row].height = 18
                current_row += 1
        else:
            ws.merge_cells(f"A{current_row}:M{current_row}")
            cell = ws.cell(row=current_row, column=1, value="(geen werkzaamheden ingevoerd)")
            cell.fill = _header_fill(LIGHT_COLOR)
            cell.font = Font(name="Segoe UI", italic=True, color="6C757D", size=10)
            cell.alignment = _left()
            cell.border = _border_thin()
            for c in range(2, 14):
                ws.cell(row=current_row, column=c).fill = _header_fill(LIGHT_COLOR)
                ws.cell(row=current_row, column=c).border = _border_thin()
            ws.row_dimensions[current_row].height = 18
            current_row += 1

    # (sheet wordt opgeslagen door de aanroeper)


def export_week_to_excel(data: dict, filepath: Path) -> None:
    """Exporteert één week naar een nieuw Excel-bestand."""
    wb = Workbook()
    ws = wb.active
    week_nr = data.get("project_week_nr", "")
    kalender_week = data.get("kalender_week_nr", "")
    ws.title = f"Week {week_nr} (KW{kalender_week})"
    _write_week_to_sheet(ws, data)
    wb.save(filepath)


def export_weeks_to_excel(weeks: list[dict], filepath: Path) -> None:
    """Exporteert meerdere weken naar één Excel-bestand, elke week op een eigen tabblad."""
    wb = Workbook()
    # Verwijder het standaard lege blad
    wb.remove(wb.active)

    # Sorteer chronologisch op jaar + kalenderweek
    weken_gesorteerd = sorted(weeks, key=lambda w: (w.get("iso_year", 0), w.get("kalender_week_nr", 0)))

    for data in weken_gesorteerd:
        iso_year = data.get("iso_year", "")
        kalender_week = data.get("kalender_week_nr", "?")
        week_start = data.get("week_start", "")
        try:
            from datetime import date as _date
            dt = _date.fromisoformat(week_start)
            datum_kort = dt.strftime("%d-%m")
        except Exception:
            datum_kort = ""
        if datum_kort:
            tab_naam = f"KW{kalender_week:02d} ({datum_kort})" if isinstance(kalender_week, int) else f"KW{kalender_week} ({datum_kort})"
        else:
            tab_naam = f"KW{kalender_week:02d} {iso_year}" if isinstance(kalender_week, int) else f"KW{kalender_week}"
        # Excel tabblad-namen max 31 tekens
        tab_naam = tab_naam[:31]
        ws = wb.create_sheet(title=tab_naam)
        _write_week_to_sheet(ws, data)

    wb.save(filepath)
